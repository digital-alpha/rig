from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework import status
from .serializers import DocumentSerializer, DetailSerializer,RoleSerializer
from rest_framework.decorators import api_view, permission_classes,renderer_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import StaticHTMLRenderer



from django.shortcuts import render,redirect
from django.http import JsonResponse
from django.views import View
import time
from UploadMulti.forms import DocumentForm
from UploadMulti.models import Document, Detail,Role
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views import View
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.contrib import messages

from django.http import HttpResponse
from django.http import HttpResponseRedirect

from UploadMulti.forms import DetailForm
import time
from Source.Entity import Entities,display_attributes
import spacy
from spacy import displacy
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font


import requests

from django.template.defaulttags import register
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

wb =Workbook()
filepath=os.path.join(BASE_DIR,'static/Excel_Files/Features.xlsx')
wb.save(filepath)

from UploadMulti.serializers import UserSerializer



nlp = spacy.load('sample_work_model_300_drop_0.05')
nlp2 = spacy.load("en_core_web_sm")


@api_view(['GET'])
def current_user(request):
    """
    Determine the current user by their token, and return their data
    """
    
    serializer = UserSerializer(request.user)
    return Response(serializer.data)

class DocumentViewset(viewsets.ModelViewSet):
    
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]


from django.forms.models import model_to_dict
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def processAPI(request):
   
    documents = Document.objects.all()
    processed_date = datetime.now()

    for document in documents:

        print(document.processed_date, processed_date)
        if Detail.objects.filter(doc_id=document.id).exists()==False:
            document.processed_date = processed_date
            document.save()
            tup=[]
            try:
                id=Detail.objects.latest('id')
                id=model_to_dict(id)['id']
            except:
                id=0

            tup.append(int(id) + 1)
            tup.append(document.file.name)
            with open('media/'+document.file.name, 'r', encoding='UTF-8') as f:
                data2=f.read()
            data2 = data2.lstrip()
            data2 = data2.rstrip()
            doc=nlp(data2)

            obj = Entities()
            mapping = obj.results(doc)
            df = obj.results_to_df(mapping)
            entities=df.to_dict('dict')
            
            # role insertion

            role=entities['Role'][0]
          
            try:
                r=Role(Role_Name=role)
                r.save()
            except:
                print("Role already exist")

            role_queryset=Role.objects.filter(Role_Name=role)
            
            role_id=list(role_queryset.values('id'))

            
            # role insertion
            for key,value in entities.items():
                tup.append(value[0])

            
            tup.append(document.id)
            tup.append(role_id[0]['id'])
            tup=tuple(tup)
            d=Detail(*tup)
            d.save()
            
            print("IN api view")
    return Response({'processed_date': processed_date}, status=200)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def processApiSingle(request, pk):
    document = get_object_or_404(Document, id=pk)
    processed_date = datetime.now()
    tup=[]
    try:
        id=Detail.objects.latest('id')
        id=model_to_dict(id)['id']
    except:
        id=0

    tup.append(int(id) + 1)
    tup.append(document.file.name)
    with open('media/'+document.file.name, 'r', encoding='UTF-8') as f:
        data2=f.read()
    data2 = data2.lstrip()
    data2 = data2.rstrip()
    doc=nlp(data2)

    obj = Entities()
    mapping = obj.results(doc)
    df = obj.results_to_df(mapping)
    entities=df.to_dict('dict')
    

    for j in entities.values():
        tup.append(j[0])

    tup.append(document.id)
    print(tup)
    tup=tuple(tup)
    d=Detail(*tup)
    d.save()
    
    print("In single api view")
    return Response({'processed_date':processed_date}, status=200)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def clearAPI(request):
    print(request.method)
    for document in Document.objects.all():
        document.file.delete()
        document.delete()
    for detail in Detail.objects.all():
        detail.delete()
    return Response(status=200)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def excelAPI(request):
    book = load_workbook(filepath)
    list_values=[]
    
    val=Detail.objects.all().values()
    field=Detail._meta.get_fields()
    field=[f.name for f in field]
    list_values.append(tuple(field[1:-1]))


    for v in val:
        value_list= [entry for entry in v.values()]
        list_values.append(tuple(value_list[1:-1]))

    writer = pd.ExcelWriter(filepath, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    df=pd.DataFrame(list(list_values))
    #print(df)
    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname,  index = False,header= False)

    ws = book[sheetname]
    black_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = black_font

    writer.save()
    if os.path.exists(filepath):
        with open(filepath, 'rb') as fh:
            response = HttpResponse(fh, content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(filepath)
            return response



def clearSingleApi(request, pk):
    print("in clearSingleApi")
    doc = get_object_or_404(Document, id=pk)
    doc.delete()
    return Response(status=200)


@api_view()
@permission_classes([IsAuthenticated])
def infoAPI(request,pk):
    
    if request.method == "GET":
        try:
            info=Detail.objects.filter(doc_id=pk)

        except Detail.DoesNotExist:

            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = DetailSerializer(info, many=True)
        return Response(serializer.data)
    if request.method == "DELETE":
        try:
            info=Detail.objects.filter(doc_id=pk)

        except Detail.DoesNotExist:

            return Response(status=status.HTTP_404_NOT_FOUND)

        info.delete()

        return Response({"message": "details deleted"}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@renderer_classes([StaticHTMLRenderer])
def DisplacyAPI(request, pk):
    doc_obj = get_object_or_404(Document, pk=pk)
    tup=[]
    form_ent=[]
    form_ent.append(doc_obj.file.name)
    with open('media/'+doc_obj.file.name, 'r', encoding='UTF-8') as f:
        data2=f.read()
    data2 = data2.lstrip()
    data2 = data2.rstrip()
    doc=nlp(data2)
    obj = Entities()
    mapping = obj.results(doc)
    d=display_attributes()
    color_scheme=d.color_table(list(mapping.keys()))
    color=[]
    color.append('#ffffff')
    color.extend(list(color_scheme.values()))
    html = displacy.render(doc, style="ent", page=True,options=d.color_dict())
    return Response(html)


class DetailViewset(viewsets.ModelViewSet):
    
    queryset = Detail.objects.all()
    serializer_class = DetailSerializer

    permission_classes = [IsAuthenticated]

class RoleViewset(viewsets.ModelViewSet):
 
    queryset = Role.objects.all()
    serializer_class = RoleSerializer

    permission_classes = [IsAuthenticated]
