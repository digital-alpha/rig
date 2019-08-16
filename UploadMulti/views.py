from django.shortcuts import render,redirect
from django.http import JsonResponse
from django.views import View
import time
from .forms import DocumentForm
from .models import Document, Detail,Role
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views import View
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.contrib import messages

from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.contrib.auth.mixins import LoginRequiredMixin

from .forms import DetailForm
import time
from Source.Entity import Entities,display_attributes
import spacy
from spacy import displacy
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from UploadMulti.API.api_views import *


import requests

from django.template.defaulttags import register
from rest_framework.authtoken.models import Token

import os
import pandas
from openpyxl import load_workbook

nlp = spacy.load('sample_work_model_300_drop_0.05')
nlp2 = spacy.load("en_core_web_sm")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
print(BASE_DIR)

list_values=[]
files=[]
# Create your views here.


wb =Workbook()

filepath=os.path.join(BASE_DIR,'UploadMulti/static/Excel_Files/Features.xlsx')
#sheet=wb.active
#sheet.append(('Employee Name', 'Address of Employee', 'Company Name', 'Address of Company', 'Role', 'Base Salary', 'Date of Agreement', 'Start Date', 'End Date', 'Supervisor Information', 'Bonus', 'Notice Period', 'Other Compensation', 'Non Monetary Benefits', 'Health Insurance', '401k', 'At will', 'Stock', 'Vacation'))
wb.save(filepath)

class BasicUploadView(LoginRequiredMixin,View):
    def get(self, request):
        documents_list = Document.objects.all()
        return render(self.request, 'UploadMulti/basic_upload/index.html', {'documents': documents_list})
        # return redirect('UploadMulti:basic_upload', {'documents': documents_list})

    def post(self, request):
        form = DocumentForm(self.request.POST, self.request.FILES)
        if form.is_valid():
            document = form.save()
            print(document.id)
            name = document.file.name
            name = name.replace('docs/', '')
            name = name.replace('.txt', '')
            date = document.uploaded_at
            print(date)
            print(name)
            data = {'is_valid': True, 'name': name, 'uploaded_at': date}
        else:
            data = {'is_valid': False}
        return JsonResponse(data)
        # documents_list = Document.objects.all()
        # return render(self.request, 'UploadMulti/basic_upload/index.html', {'documents': documents_list})

class DragAndDropUploadView(View):
    def get(self, request):
        documents_list = Document.objects.all()
        return render(self.request, 'UploadMulti/drag_and_drop_upload/index.html', {'documents': documents_list})

    def post(self, request):
        form = DocumentForm(self.request.POST, self.request.FILES)
        if form.is_valid():
            document = form.save()
            data = {'is_valid': True, 'name': document.file.name, 'url': document.file.url}
        else:
            data = {'is_valid': False}
        return JsonResponse(data)


def clear_database(request):
    for document in Document.objects.all():
        document.file.delete()
        document.delete()
        wb=Workbook()
        wb.save(filepath)
    for detail in Detail.objects.all():
        detail.delete()
    return redirect(request.POST.get('next'))

def analysis(request, pk):
    doc_obj = get_object_or_404(Document, pk=pk)
    
    tup=[]
    form_ent=[]
    form_ent.append(doc_obj.file.name)
    with open('media/'+doc_obj.file.name, 'r', encoding='UTF-8') as f:
        data2=f.read()
    data2 = data2.lstrip()
    data2 = data2.rstrip()
    doc=nlp(data2)

    obj = Entities()
    mapping = obj.results(doc)

    if doc.ents:
        for ent in doc.ents:
            if ent.label_=="Date_aggrement":
                print(type(ent.label))
                print(type(ent))
                
    
    d=Detail.objects.select_related().filter(Document_Name=doc_obj.file.name).values()
    
    list_result = [entry for entry in d]
    field=Detail._meta.get_fields()

    # getting correponding role before displaying
    #role_query=Role.objects.filter(id=list_result[0]['Role_id'])
    #role=list(role_query.values('Role_Name'))
    #list_result[0]['Role_id']=role[0]['Role_Name']
    print(len(list_result))
    value_dict=list_result[0]
    value_list=list(value_dict.values())
    value_list=value_list[1:-2]
    print(value_list)
    print(len(value_list))
    tup=value_list
    
    d=display_attributes()
    color_scheme=d.color_table(list(mapping.keys()))
    color=[]
    color.append('#ffffff')
    color.extend(list(color_scheme.values()))
   # print(entities.keys())
   

    html = displacy.render(doc, style="ent", page=True,options=d.color_dict())
    DISPLACY_DIR=os.path.join(BASE_DIR,'/UploadMulti/static/Displacy_html')
    print(DISPLACY_DIR +'/doc{}.html')
    with open(BASE_DIR+'/UploadMulti/static' +'/doc{}.html'.format(pk), 'w') as myfile:
        myfile.write(html)


    # if this doesnt work add the file and proceed.SSS
    return render(request, 'UploadMulti/analysis.html', context={'File_Name':'doc{}.html'.format(pk),'color':color_scheme,'pk':pk,'form':DetailForm(dynamic_placeholder=tup,doc_color=color)})

    # if this doesnt work add the file and proceed.
    #


   # return render(request, 'analysis.html', {'doc_obj':doc_obj})


...
@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)


def csv(request):

    book = load_workbook(filepath)
    list_values=[]
    
    val=Detail.objects.all().values()
    field=Detail._meta.get_fields()
    field=[f.name for f in field]
    list_values.append(tuple(field[1:-1]))


    for v in val:
        value_list= [entry for entry in v.values()]
        list_values.append(tuple(value_list[1:-1]))

    writer = pd.ExcelWriter(filepath, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    df=pd.DataFrame(list(list_values))
    #print(df)
    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname,  index = False,header= False)

    ws = book[sheetname]
    black_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = black_font

    writer.save()
    if os.path.exists(filepath):
        with open(filepath, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(filepath)
            return response

    raise Http404
   


import requests
def process(request):
    documents = Document.objects.all()
    tokens=Token.objects.filter(user=request.user).values()
    list_result = [entry for entry in tokens]
    token=list_result[0]['key']
    myurl = "http://localhost:8000/api/process/"
    response = requests.get(myurl, headers={'Authorization': 'Token {}'.format(token)})
   
    print(response)
    #if request.method == 'GET':

    #    p=processAPI(request)  
    #    print(p)


    return render(request, 'UploadMulti/basic_upload/index.html', {'documents':documents})

def clearSingleDoc(request, pk):
    print("In clearSIngleDoc view")
    tokens=Token.objects.filter(user=request.user).values()
    list_result = [entry for entry in tokens]
    token=list_result[0]['key']
    myurl = "http://localhost:8000/api/clear/" + str(pk)
    response = requests.get(myurl, headers={'Authorization': 'Token {}'.format(token)}, )
    documents = Document.objects.all()
    return render(request, 'UploadMulti/basic_upload/index.html', {'documents':documents})

def form_post(request):

    if request.method == "POST":

        p = list(request.POST.values())

        instance=get_object_or_404(Detail,Document_Name=p[1])
        p = p[1:]
        print(p)
        field=Detail._meta.get_fields()
        field=[f.name for f in field ]
        print(field)
        field.remove('doc')
        field.remove('Role_ref')
        
       
        dict_info = dict(zip(field[1:], p))
        #print(dict_info)
        
        
        # getting the corresponding role_id before insertion 
        try:
            role_query=Role.objects.filter(Role_Name=dict_info['Roles'])
            role_id=list(role_query.values('id'))
            role_id=role_id[0]['id']
        except:
            r=Role(Role_Name=dict_info['Roles'])
            r.save()
            role_query=Role.objects.filter(Role_Name=dict_info['Roles'])
            role_id=list(role_query.values('id'))
            role_id=role_id[0]['id']



        

        p=list(dict_info.values())
        #print(p)

        
       
        form = DetailForm(request.POST,dynamic_placeholder=p,instance=instance)
      
       
        if form.is_valid():
            print("Form save")
            p=form.save()
            p.Role_ref_id=role_query=Role.objects.get(Role_Name=dict_info['Roles']).id
            p.save()
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
          
        else:
            messages.error(request, 'The form is invalid.')
            print(form.errors)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))



# from dal import autocomplete



# class CountryAutocomplete(autocomplete.Select2ListView):

#     def create(self, text):
#         return text

#     def get_list(self):
#         result_list = []
#         if self.q:
#             data = Detail.objects.all().filter(Role__icontains=self.q)[:10]

#             result_list = [x.name for x in data]
#             print(result_list)
#             print("Herer")
#         return result_list


########################################################################


from dal import autocomplete
#from your_countries_app.models import Country


class ClientAutocomplete(autocomplete.Select2QuerySetView):
    def get(self):
        # Don't forget to filter out results depending on the visitor !
        print("here")
        qs = Role.objects.all()

        if self.q:
            
            qs = qs.filter(Role_Name__istartswith=self.q)
            print(qs)

        return qs